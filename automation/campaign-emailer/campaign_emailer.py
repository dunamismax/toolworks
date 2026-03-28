#!/usr/bin/env python3
"""
campaign-emailer: Send templated marketing emails via Outlook on macOS.

Reads an Excel campaign tracker, finds the next batch of un-emailed contacts,
fills in the email template (HTML with signature), and sends via AppleScript-driven Outlook.

Usage:
    python3 campaign_emailer.py                  # send next 10
    python3 campaign_emailer.py --count 5        # send next 5
    python3 campaign_emailer.py --dry-run        # preview without sending
    python3 campaign_emailer.py --status         # show tracker status
    python3 campaign_emailer.py --test           # send one test to personal email (no CC)
"""

from __future__ import annotations

import argparse
import os
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl

# --- Configuration ---

DEFAULT_TRACKER_PATH = (
    "/Users/sawyer/Library/CloudStorage/OneDrive-ImagingServicesInc/"
    "OneDrive - Work Desktop/07-Marketing/Campaign-Tracker/"
    "contacts-campaign-tracker.xlsx"
)
DEFAULT_TEMPLATE_PATH = str(Path(__file__).parent / "templates" / "v2-coverage-offer.html")
DEFAULT_CC_EMAIL = "support@imagingservices.net"
DEFAULT_SUBJECT = "Your imaging system's support coverage has lapsed"
DEFAULT_TEST_EMAIL = "stephenvsawyer@gmail.com"

TRACKER_PATH = os.environ.get("CAMPAIGN_EMAILER_TRACKER_PATH", DEFAULT_TRACKER_PATH)
TEMPLATE_PATH = os.environ.get("CAMPAIGN_EMAILER_TEMPLATE_PATH", DEFAULT_TEMPLATE_PATH)
CC_EMAIL = os.environ.get("CAMPAIGN_EMAILER_CC_EMAIL", DEFAULT_CC_EMAIL)
SUBJECT = os.environ.get("CAMPAIGN_EMAILER_SUBJECT", DEFAULT_SUBJECT)
TEST_EMAIL = os.environ.get("CAMPAIGN_EMAILER_TEST_EMAIL", DEFAULT_TEST_EMAIL)

# Column indices (0-based)
COL_COMPANY = 1  # Company Name
COL_CONTACT = 2  # Contact Name
COL_EMAIL = 4  # Email Address
COL_EMAILED = 6  # Emailed (TRUE/FALSE)
COL_DATE = 7  # Date Emailed
COL_STATUS = 15  # Status

# Contact names to skip (use company name instead)
SKIP_NAMES = {
    "reception",
    "receptionist",
    "front desk",
    "doctor",
    "dr.",
    "dr",
    "office",
    "admin",
    "manager",
    "staff",
    "team",
    "billing",
    "office manager",
    "front office",
    "n/a",
    "na",
    "none",
    "",
    "parent billing company",
    "billing company",
    "owner",
    "unknown",
    "tech",
    "technician",
    "technologist",
    "practice",
    "clinic",
    "hospital",
    "center",
    "llc",
    "inc",
    "associates",
    "group",
    "contact",
    "info",
    "general",
    "main",
    "support",
    "service",
    "services",
    "accounts",
    "accounting",
    "payable",
    "receivable",
}

# Company names that are not real practice names (fall back to generic greeting)
SKIP_COMPANIES = {
    "parent billing company",
    "billing company",
    "unknown",
    "n/a",
    "na",
    "none",
    "test",
    "",
}

# Words in a contact name that suggest it's a business name, not a person
BUSINESS_NAME_MARKERS = {
    "llc",
    "inc",
    "corp",
    "corporation",
    "associates",
    "partners",
    "veterinary",
    "chiropractic",
    "podiatry",
    "orthopedic",
    "ortho",
    "hospital",
    "clinic",
    "center",
    "practice",
    "medical",
    "health",
    "animal",
    "foot",
    "ankle",
    "spine",
    "family",
    "group",
    "services",
}

TRUTHY_FLAGS = {"true", "yes", "y", "1"}
FALSEY_FLAGS = {"false", "no", "n", "0", ""}


@dataclass(frozen=True)
class PendingContact:
    row_num: int
    company: str
    contact: str | None
    email: str


def build_tracker_lock_path(tracker_path: str) -> str:
    """Return the Excel lockfile path for the tracker workbook."""
    tracker = Path(tracker_path)
    return str(tracker.with_name(f"~${tracker.name}"))


TRACKER_LOCK = build_tracker_lock_path(TRACKER_PATH)


def clean_name(value: str | None) -> str:
    """Remove parenthetical notes and normalize surrounding whitespace."""
    if not value:
        return ""
    cleaned = re.sub(r"\s*\([^)]*\)\s*", " ", value)
    return " ".join(cleaned.split())


def normalize_contact_name(name: str) -> str:
    """Normalize obvious casing issues in contact names for greetings."""
    cleaned = clean_name(name)
    if not cleaned:
        return ""

    parts = []
    for token in cleaned.split():
        letters_only = re.sub(r"[^A-Za-z]", "", token)
        if letters_only and (letters_only.isupper() or letters_only.islower()):
            token = token.title()
        parts.append(token)

    return " ".join(parts)


def is_bad_greeting_token(token: str) -> bool:
    """Check whether a greeting token is too generic or malformed to use."""
    normalized = re.sub(r"[^A-Za-z]", "", token).lower()
    if not normalized:
        return True
    if len(normalized) == 1:
        return True
    if normalized in SKIP_NAMES:
        return True
    if normalized in BUSINESS_NAME_MARKERS:
        return True
    return False


def is_generic_name(name: str | None) -> bool:
    """Check if a contact name is generic and should be replaced."""
    cleaned = clean_name(name).lower()
    if cleaned in SKIP_NAMES:
        return True
    words = set(cleaned.split())
    return bool(words & BUSINESS_NAME_MARKERS)


def is_generic_company(name: str | None) -> bool:
    """Check if a company name is too generic to use in a greeting."""
    cleaned = clean_name(name).lower()
    return cleaned in SKIP_COMPANIES


def build_greeting(contact_name: str | None, company_name: str) -> str:
    """Build a natural greeting line from contact and company data."""
    if contact_name and not is_generic_name(contact_name):
        clean = normalize_contact_name(contact_name)
        parts = clean.split()

        if parts:
            first = re.sub(r"^[^A-Za-z]+|[^A-Za-z]+$", "", parts[0])
            if first.lower().rstrip(".") == "dr" and len(parts) > 1:
                doctor_name = re.sub(r"^[^A-Za-z]+|[^A-Za-z]+$", "", parts[1])
                if doctor_name and not is_bad_greeting_token(doctor_name):
                    return f"Hi Dr. {doctor_name},"
                return "Hi there,"

            if first and not is_bad_greeting_token(first):
                return f"Hi {first},"

    if not is_generic_company(company_name):
        return f"Hi {clean_name(company_name)} team,"

    return "Hi there,"


def load_template() -> str:
    """Load the HTML template file."""
    with open(TEMPLATE_PATH, encoding="utf-8") as handle:
        return handle.read()


def personalize_html(html: str, contact_name: str | None, company_name: str) -> str:
    """Replace greeting placeholder in the HTML template."""
    greeting = build_greeting(contact_name, company_name)
    return html.replace("{{GREETING}}", greeting)


def applescript_quote(value: str) -> str:
    """Escape a Python string for safe use inside an AppleScript string literal."""
    return value.replace("\\", "\\\\").replace('"', '\\"')


def send_html_via_applescript(
    to_email: str,
    cc_email: str | None,
    subject: str,
    html_body: str,
) -> bool:
    """Send an HTML email via Outlook using AppleScript with a temp HTML file."""
    with tempfile.NamedTemporaryFile(
        mode="w",
        suffix=".html",
        delete=False,
        encoding="utf-8",
    ) as handle:
        handle.write(html_body)
        html_path = handle.name

    quoted_path = applescript_quote(html_path)
    quoted_subject = applescript_quote(subject)
    quoted_to_email = applescript_quote(to_email)

    try:
        if cc_email:
            cc_line = (
                "make new cc recipient at newMessage with properties "
                f'{{email address:{{address:"{applescript_quote(cc_email)}"}}}}'
            )
        else:
            cc_line = ""

        message_line = (
            "set newMessage to make new outgoing message with properties "
            f'{{subject:"{quoted_subject}", content:htmlContent}}'
        )
        recipient_line = (
            "make new to recipient at newMessage with properties "
            f'{{email address:{{address:"{quoted_to_email}"}}}}'
        )

        script = f'''
        set htmlFile to POSIX file "{quoted_path}"
        set htmlContent to read htmlFile as «class utf8»

        tell application "Microsoft Outlook"
            {message_line}
            {recipient_line}
            {cc_line}
            send newMessage
        end tell
        '''

        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if result.returncode != 0:
            print(f"  AppleScript error: {result.stderr.strip()}", file=sys.stderr)
            return False
        return True
    except subprocess.TimeoutExpired:
        print("  AppleScript timed out", file=sys.stderr)
        return False
    except Exception as exc:
        print(f"  AppleScript exception: {exc}", file=sys.stderr)
        return False
    finally:
        Path(html_path).unlink(missing_ok=True)


def flag_is_truthy(value: object) -> bool:
    """Normalize workbook truthy flag values."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in TRUTHY_FLAGS
    if isinstance(value, (int, float)):
        return value == 1
    return False


def flag_is_falsey(value: object) -> bool:
    """Normalize workbook falsey flag values."""
    if value is None:
        return True
    if isinstance(value, bool):
        return not value
    if isinstance(value, str):
        return value.strip().lower() in FALSEY_FLAGS
    if isinstance(value, (int, float)):
        return value == 0
    return False


def get_pending_rows(ws, max_count: int) -> list[PendingContact]:
    """Find the next batch of un-emailed contacts."""
    rows: list[PendingContact] = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        emailed = vals[COL_EMAILED]

        if flag_is_falsey(emailed):
            email_addr = vals[COL_EMAIL]
            if not email_addr or not isinstance(email_addr, str) or "@" not in email_addr:
                continue

            rows.append(
                PendingContact(
                    row_num=row[0].row,
                    company=vals[COL_COMPANY] or "Unknown Company",
                    contact=vals[COL_CONTACT],
                    email=email_addr.strip(),
                )
            )

            if len(rows) >= max_count:
                break
    return rows


def update_tracker(ws, row_num: int) -> None:
    """Mark a row as emailed in the tracker."""
    ws.cell(row=row_num, column=COL_EMAILED + 1, value=True)
    ws.cell(row=row_num, column=COL_DATE + 1, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=row_num, column=COL_STATUS + 1, value="In Progress")


def load_workbook_or_exit(*, read_only: bool):
    """Load the tracker workbook or exit with a clear error."""
    try:
        return openpyxl.load_workbook(TRACKER_PATH, read_only=read_only)
    except Exception as exc:  # pragma: no cover - depends on local filesystem/workbook state
        print(f"ERROR: Unable to open tracker workbook: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc


def show_status() -> None:
    """Show current tracker status."""
    wb = load_workbook_or_exit(read_only=True)
    ws = wb.active

    total = 0
    emailed = 0
    remaining = 0
    next_row = None

    for row in ws.iter_rows(min_row=2, values_only=False):
        vals = [cell.value for cell in row]
        if vals[COL_EMAIL]:
            total += 1
            if flag_is_truthy(vals[COL_EMAILED]):
                emailed += 1
            else:
                remaining += 1
                if next_row is None:
                    next_row = {
                        "row": row[0].row,
                        "company": vals[COL_COMPANY],
                        "contact": vals[COL_CONTACT],
                        "email": vals[COL_EMAIL],
                    }

    wb.close()

    print("Campaign Tracker Status")
    print(f"  Total contacts:    {total}")
    print(f"  Emailed:           {emailed}")
    print(f"  Remaining:         {remaining}")
    if next_row:
        print(
            f"  Next up (row {next_row['row']}): "
            f"{next_row['contact'] or next_row['company']} <{next_row['email']}>"
        )


def send_test(html_template: str) -> None:
    """Send a test email to personal address. No CC, no tracker update."""
    wb = load_workbook_or_exit(read_only=True)
    ws = wb.active

    pending = get_pending_rows(ws, 1)
    wb.close()

    if not pending:
        print("No pending contacts to use as test data.")
        return

    contact = pending[0]
    greeting = build_greeting(contact.contact, contact.company)
    html = personalize_html(html_template, contact.contact, contact.company)

    print(f"Sending TEST email to {TEST_EMAIL}")
    print(f"  Subject: [TEST] {SUBJECT}")
    print(f"  Using data from row {contact.row_num}: {contact.company}")
    print(f"  Contact: {contact.contact or '(none)'}")
    print(f"  Greeting: {greeting}")
    print("  CC: (none - test mode)")

    success = send_html_via_applescript(TEST_EMAIL, None, f"[TEST] {SUBJECT}", html)
    if success:
        print(f"  -> Test sent to {TEST_EMAIL}")
    else:
        print("  -> Test FAILED")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Send campaign emails via Outlook")
    parser.add_argument(
        "--count",
        "-n",
        type=int,
        default=10,
        help="Number of emails to send (default: 10)",
    )
    parser.add_argument(
        "--dry-run",
        "-d",
        action="store_true",
        help="Preview emails without sending",
    )
    parser.add_argument(
        "--status",
        "-s",
        action="store_true",
        help="Show tracker status and exit",
    )
    parser.add_argument(
        "--test",
        "-t",
        action="store_true",
        help="Send one test to personal email (no CC)",
    )
    args = parser.parse_args(argv)

    if args.count < 1:
        parser.error("--count must be at least 1")

    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if args.status:
        show_status()
        return 0

    if not os.path.exists(TRACKER_PATH):
        print(f"ERROR: Tracker not found: {TRACKER_PATH}", file=sys.stderr)
        return 1

    if not os.path.exists(TEMPLATE_PATH):
        print(f"ERROR: Template not found: {TEMPLATE_PATH}", file=sys.stderr)
        return 1

    # Check if Excel has the tracker open
    if not args.dry_run and not args.test and os.path.exists(TRACKER_LOCK):
        print("WARNING: The campaign tracker is open in Excel.")
        print("   The tracker CANNOT be updated while Excel has it open.")
        print("   Close the file in Excel first, then run again.")
        print(f"   Lock file: {TRACKER_LOCK}")
        return 1

    html_template = load_template()

    if args.test:
        send_test(html_template)
        return 0

    read_only = args.dry_run
    wb = load_workbook_or_exit(read_only=read_only)
    ws = wb.active

    pending = get_pending_rows(ws, args.count)
    if not pending:
        print("No pending contacts found. All caught up!")
        wb.close()
        return 0

    print(f"{'[DRY RUN] ' if args.dry_run else ''}Processing {len(pending)} email(s)...\n")

    sent = 0
    failed = 0

    for contact in pending:
        html = personalize_html(html_template, contact.contact, contact.company)
        greeting = build_greeting(contact.contact, contact.company)

        print(f"  Row {contact.row_num}: {contact.company}")
        print(f"    To: {contact.email}")
        print(f"    Greeting: {greeting}")

        if args.dry_run:
            print("    -> [DRY RUN] Would send")
            sent += 1
        else:
            success = send_html_via_applescript(contact.email, CC_EMAIL, SUBJECT, html)
            if success:
                update_tracker(ws, contact.row_num)
                print("    -> Sent")
                sent += 1
            else:
                print("    -> FAILED")
                failed += 1

        print()

    if not args.dry_run and sent > 0:
        wb.save(TRACKER_PATH)
        print(f"Tracker updated ({sent} rows marked TRUE).")

    wb.close()

    print(f"\nDone. Sent: {sent}, Failed: {failed}")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
